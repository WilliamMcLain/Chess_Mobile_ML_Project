"""
computer.py — GNN chess computer with score.xlsx integration.

score.xlsx columns:
  A: game number   B: # of moves   C: winning side   D+: each move as "pt start end"

GNN trains from score.xlsx on every game end and on startup (pretrain).
"""

import os, sys, random
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import pieces, csv_setup

try:
    import torch, torch.nn as nn, torch.nn.functional as F
    from torch_geometric.nn import GCNConv, global_mean_pool
    from torch_geometric.data import Data
    TORCH_AVAILABLE = True
except ImportError:
    TORCH_AVAILABLE = False

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

PIECE_VALUES = {'p':1,'h':3,'b':3,'c':5,'q':9,'k':100}
PIECE_IDX    = {'p':0,'h':1,'b':2,'c':3,'q':4,'k':5,'e':6}
DATA_DIR     = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
MODEL_PATH   = os.path.join(DATA_DIR, 'chess_gnn.pt')
SCORE_PATH   = os.path.join(DATA_DIR, 'score.xlsx')


# ─── score.xlsx interface ─────────────────────────────────────────────────────

def _ensure_score_file():
    if not XLSX_AVAILABLE:
        return
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(SCORE_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['game (number','#of moves','winning side','move 1','move 2','move 3'])
        wb.save(SCORE_PATH)

def score_next_game_number() -> int:
    if not XLSX_AVAILABLE or not os.path.exists(SCORE_PATH):
        return 1
    wb = openpyxl.load_workbook(SCORE_PATH, data_only=True)
    return wb.active.max_row  # row 1 = header, so max_row = game count + 1

def score_write_game(game_number: int, moves: list, winning_side: str):
    if not XLSX_AVAILABLE:
        return
    _ensure_score_file()
    wb = openpyxl.load_workbook(SCORE_PATH)
    ws = wb.active
    # Extend header row if this game has more moves than current columns
    needed = 3 + len(moves)
    current_cols = ws.max_column
    for i in range(current_cols, needed):
        ws.cell(row=1, column=i+1, value=f'move {i-2}')
    ws.append([game_number, len(moves), winning_side] + moves)
    wb.save(SCORE_PATH)
    print(f"[score] Game {game_number} written → score.xlsx  winner={winning_side}  moves={len(moves)}")

def score_read_all_games() -> list:
    if not XLSX_AVAILABLE or not os.path.exists(SCORE_PATH):
        return []
    wb  = openpyxl.load_workbook(SCORE_PATH, data_only=True)
    ws  = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        out.append({
            'game':      int(row[0]),
            'moves':     int(row[1]) if row[1] else 0,
            'winner':    row[2],
            'move_list': [m for m in row[3:] if m is not None]
        })
    return out

def score_move_frequency() -> dict:
    freq = {}
    for g in score_read_all_games():
        for i, move in enumerate(g['move_list']):
            if move not in freq:
                freq[move] = {'count': 0, 'wins': 0}
            freq[move]['count'] += 1
            mover = 'white' if i % 2 == 0 else 'black'
            if mover == g['winner']:
                freq[move]['wins'] += 1
    for m in freq:
        c = freq[m]['count']
        freq[m]['win_rate'] = freq[m]['wins'] / c if c else 0.0
    return freq

def score_get_training_pairs() -> list:
    pairs = []
    for g in score_read_all_games():
        for i, move in enumerate(g['move_list']):
            mover  = 'white' if i % 2 == 0 else 'black'
            reward = 1.0 if mover == g['winner'] else -1.0
            pairs.append((move, reward))
    return pairs

def score_print_summary():
    games = score_read_all_games()
    if not games:
        print("[score] No games recorded yet.")
        return
    total  = len(games)
    whites = sum(1 for g in games if g['winner'] == 'white')
    blacks = sum(1 for g in games if g['winner'] == 'black')
    avg_m  = sum(g['moves'] for g in games) / total
    print(f"[score] Total games: {total}  White wins: {whites}  "
          f"Black wins: {blacks}  Draws: {total-whites-blacks}  Avg moves: {avg_m:.1f}")


# ─── GNN ─────────────────────────────────────────────────────────────────────

class ChessGNN(nn.Module):
    def __init__(self):
        super().__init__()
        self.conv1 = GCNConv(16, 64);   self.bn1 = nn.BatchNorm1d(64)
        self.conv2 = GCNConv(64, 128);  self.bn2 = nn.BatchNorm1d(128)
        self.conv3 = GCNConv(128, 256); self.bn3 = nn.BatchNorm1d(256)
        self.conv4 = GCNConv(256, 128); self.bn4 = nn.BatchNorm1d(128)
        self.conv5 = GCNConv(128, 64);  self.bn5 = nn.BatchNorm1d(64)
        self.fc1   = nn.Linear(64, 32); self.fc2 = nn.Linear(32, 1)
        self.drop  = nn.Dropout(0.2)

    def forward(self, data):
        x, ei = data.x, data.edge_index
        batch  = data.batch if hasattr(data,'batch') and data.batch is not None \
                 else torch.zeros(x.size(0), dtype=torch.long)
        x = self.drop(F.relu(self.bn1(self.conv1(x, ei))))
        x = self.drop(F.relu(self.bn2(self.conv2(x, ei))))
        x = self.drop(F.relu(self.bn3(self.conv3(x, ei))))
        x = F.relu(self.bn4(self.conv4(x, ei)))
        x = F.relu(self.bn5(self.conv5(x, ei)))
        x = global_mean_pool(x, batch)
        return self.fc2(F.relu(self.fc1(x))).squeeze(-1)


# ─── board utilities ──────────────────────────────────────────────────────────

def read_board() -> dict:
    files = 'abcdefgh'
    return {files[fi]+str(ri+1): (csv_setup.csv_setup.find(files[fi]+str(ri+1)) or 'e')
            for ri in range(8) for fi in range(8)}

def apply_move(board: dict, pt: str, start: str, end: str, side: str) -> dict:
    nb = dict(board); nb[end] = side+pt; nb[start] = 'e'; return nb

def material_score(board: dict, side: str) -> float:
    h = 'b' if side == 'w' else 'w'
    return float(
        sum(PIECE_VALUES.get(p[1],0) for p in board.values() if p!='e' and p[0]==side) -
        sum(PIECE_VALUES.get(p[1],0) for p in board.values() if p!='e' and p[0]==h))

def print_score(board: dict, comp: str):
    h  = 'b' if comp == 'w' else 'w'
    cs = sum(PIECE_VALUES.get(p[1],0) for p in board.values() if p!='e' and p[0]==comp)
    hs = sum(PIECE_VALUES.get(p[1],0) for p in board.values() if p!='e' and p[0]==h)
    d  = cs - hs
    print(f"[score] Computer {cs}  Human {hs}  "
          f"{'WINNING' if d>0 else 'LOSING' if d<0 else 'EVEN'} ({'+' if d>=0 else ''}{d})")

def _attacks(pos: str, pt: str, side: str, board: dict) -> list:
    files = 'abcdefgh'
    fi, ri = ord(pos[0])-ord('a'), int(pos[1])-1
    enemy  = 'b' if side=='w' else 'w'
    result = []
    ib = lambda f,r: 0<=f<8 and 0<=r<8
    sq = lambda f,r: files[f]+str(r+1)
    def slide(df,dr):
        f,r = fi+df, ri+dr
        while ib(f,r):
            result.append(sq(f,r))
            if board.get(sq(f,r),'e') != 'e': break
            f+=df; r+=dr
    if pt=='p':
        d = 1 if side=='w' else -1
        for df in [-1,1]:
            if ib(fi+df, ri+d): result.append(sq(fi+df, ri+d))
    elif pt=='h':
        for df,dr in [(-2,-1),(-2,1),(-1,-2),(-1,2),(1,-2),(1,2),(2,-1),(2,1)]:
            if ib(fi+df, ri+dr): result.append(sq(fi+df, ri+dr))
    elif pt=='b': [slide(df,dr) for df,dr in [(-1,-1),(-1,1),(1,-1),(1,1)]]
    elif pt=='c': [slide(df,dr) for df,dr in [(-1,0),(1,0),(0,-1),(0,1)]]
    elif pt=='q': [slide(df,dr) for df,dr in [(-1,-1),(-1,1),(1,-1),(1,1),(-1,0),(1,0),(0,-1),(0,1)]]
    elif pt=='k':
        for df,dr in [(-1,-1),(-1,0),(-1,1),(0,-1),(0,1),(1,-1),(1,0),(1,1)]:
            if ib(fi+df, ri+dr): result.append(sq(fi+df, ri+dr))
    return [s for s in result if board.get(s,'e')=='e' or board.get(s,'e')[0]==enemy]

def legal_moves(board: dict, side: str) -> list:
    files = 'abcdefgh'
    out   = []
    for pos, piece in board.items():
        if piece=='e' or piece[0]!=side: continue
        pt = piece[1]
        for ri in range(8):
            for fi in range(8):
                end = files[fi]+str(ri+1)
                if end==pos: continue
                try:
                    if pieces.Pieces.isMoveValid(side, pt, pos, end):
                        out.append((pt, pos, end))
                except Exception: pass
    return out

def board_to_graph(board: dict, comp: str):
    files = 'abcdefgh'
    idx   = lambda p: (int(p[1])-1)*8+(ord(p[0])-ord('a'))
    itp   = lambda i: files[i%8]+str(i//8+1)
    threatened, defended = set(), set()
    for pos, piece in board.items():
        if piece=='e': continue
        for sq in _attacks(pos, piece[1], piece[0], board):
            t = board.get(sq,'e')
            if t!='e': (threatened if t[0]!=piece[0] else defended).add(sq)
    feats = []
    for i in range(64):
        pos   = itp(i); piece = board.get(pos,'e'); occ = int(piece!='e')
        pt    = piece[1] if occ else 'e'
        oh    = [0]*7; oh[PIECE_IDX.get(pt,6)] = 1
        mob   = len(_attacks(pos,pt,piece[0],board))/28.0 if occ else 0
        feats.append([occ, int(occ and piece[0]==comp), int(occ and piece[0]!=comp),
                      *oh, PIECE_VALUES.get(pt,0)/100.0,
                      (int(pos[1])-1)/7.0, (ord(pos[0])-ord('a'))/7.0,
                      int(pos in threatened), int(pos in defended), mob])
    x   = torch.tensor(feats, dtype=torch.float)
    src, dst = [], []
    for pos, piece in board.items():
        if piece=='e': continue
        s = idx(pos)
        for t in _attacks(pos, piece[1], piece[0], board):
            d = idx(t); src+=[s,d]; dst+=[d,s]
    ei = torch.tensor([src,dst],dtype=torch.long) if src else torch.zeros((2,0),dtype=torch.long)
    return Data(x=x, edge_index=ei)


# ─── Computer ─────────────────────────────────────────────────────────────────

class Computer:
    def __init__(self, side: str):
        self.side        = side
        self.human       = 'b' if side=='w' else 'w'
        self.replay      = []
        self.max_buf     = 5000
        self.batch       = 32
        self.gamma       = 0.95
        self.epsilon     = 1.0
        self.eps_min     = 0.05
        self.eps_dec     = 0.995
        self.history     = []   # (board_dict, pt, start, end)
        self.move_log    = []   # "pt start end" strings → score.xlsx
        self.game_num    = score_next_game_number()
        self._freq_cache = None

        _ensure_score_file()

        if TORCH_AVAILABLE:
            self.model = ChessGNN()
            self.opt   = torch.optim.Adam(self.model.parameters(), lr=1e-3)
            self.loss  = nn.MSELoss()
            self._load()
            self._pretrain()
        else:
            self.model = None

    def _load(self):
        if not os.path.exists(MODEL_PATH):
            print("[computer] Fresh model"); return
        try:
            ck = torch.load(MODEL_PATH, map_location='cpu')
            self.model.load_state_dict(ck['model'])
            self.opt.load_state_dict(ck['optimizer'])
            self.epsilon = ck.get('epsilon', self.epsilon)
            print(f"[computer] Model loaded eps={self.epsilon:.3f}")
        except Exception as e:
            print(f"[computer] Load failed: {e}")

    def _save(self):
        os.makedirs(DATA_DIR, exist_ok=True)
        torch.save({'model':self.model.state_dict(),
                    'optimizer':self.opt.state_dict(),
                    'epsilon':self.epsilon}, MODEL_PATH)

    def _pretrain(self):
        pairs = score_get_training_pairs()
        if not pairs: return
        board = read_board()
        for _, reward in pairs:
            try:
                self.replay.append((board_to_graph(board, self.side), reward))
            except Exception: pass
        if len(self.replay) >= self.batch:
            self._train_step()
        print(f"[computer] Pretrained on {len(pairs)} historical moves from score.xlsx")

    def _score_bonus(self, pt: str, start: str, end: str) -> float:
        if self._freq_cache is None:
            self._freq_cache = score_move_frequency()
        key = f"{pt} {start} {end}"
        entry = self._freq_cache.get(key)
        return entry['win_rate'] * 0.5 if entry else 0.0

    def _eval(self, board: dict, pt: str, start: str, end: str) -> float:
        score = material_score(board, self.side)
        if self.model is not None and TORCH_AVAILABLE:
            self.model.eval()
            with torch.no_grad():
                score += self.model(board_to_graph(board, self.side)).item()
        score += self._score_bonus(pt, start, end)
        return score

    def _pick(self, moves: list, board: dict):
        if not moves: return None
        if random.random() < self.epsilon:
            return random.choice(moves)
        best, best_s = None, float('-inf')
        for pt, start, end in moves:
            s = self._eval(apply_move(board, pt, start, end, self.side), pt, start, end)
            if s > best_s: best_s = s; best = (pt, start, end)
        return best

    def move(self) -> bool:
        board = read_board()
        moves = legal_moves(board, self.side)
        if not moves:
            print(f"[computer] No legal moves for {self.side}"); return False
        chosen = self._pick(moves, board)
        if chosen is None: return False
        pt, start, end = chosen
        self.history.append((dict(board), pt, start, end))
        self.move_log.append(f"{pt} {start} {end}")
        ok = pieces.Pieces.prime(self.side, pt, start, end)
        if ok:
            print_score(read_board(), self.side)
            print(f"[computer] {self.side.upper()}: {pt} {start}→{end}")
            self.epsilon = max(self.eps_min, self.epsilon * self.eps_dec)
            self._freq_cache = None  # invalidate cache after move
        return bool(ok)

    def record_outcome(self, won: bool, all_moves: list = None):
        """Call at game end. all_moves = full game move list (both sides)."""
        winner = ('white' if self.side=='w' else 'black') if won else \
                 ('black' if self.side=='w' else 'white')
        game_moves = all_moves if all_moves else self.move_log
        score_write_game(self.game_num, game_moves, winner)
        score_print_summary()

        if TORCH_AVAILABLE and self.history:
            r = 1.0 if won else -1.0
            for bd, pt, start, end in reversed(self.history):
                self.replay.append((board_to_graph(bd, self.side), r))
                r *= self.gamma
            if len(self.replay) > self.max_buf:
                self.replay = self.replay[-self.max_buf:]
            if len(self.replay) >= self.batch:
                self._train_step()
            self._save()

        self.history  = []
        self.move_log = []
        self.game_num += 1
        self._freq_cache = None

    def _train_step(self):
        self.model.train()
        samples = random.sample(self.replay, self.batch)
        loss_total = 0.0
        for graph, reward in samples:
            self.opt.zero_grad()
            pred = self.model(graph)
            tgt  = torch.tensor([reward], dtype=torch.float)
            loss = self.loss(pred, tgt)
            loss.backward()
            torch.nn.utils.clip_grad_norm_(self.model.parameters(), 1.0)
            self.opt.step()
            loss_total += loss.item()
        print(f"[computer] loss={loss_total/self.batch:.4f} eps={self.epsilon:.3f} buf={len(self.replay)}")

    def reward_capture(self, piece: str):
        if not self.history or not TORCH_AVAILABLE: return
        v = PIECE_VALUES.get(piece[1], 0)/100.0
        self.replay.append((board_to_graph(self.history[-1][0], self.side), v))

    def reward_lost(self, piece: str):
        if not self.history or not TORCH_AVAILABLE: return
        v = -PIECE_VALUES.get(piece[1], 0)/100.0
        self.replay.append((board_to_graph(self.history[-1][0], self.side), v))